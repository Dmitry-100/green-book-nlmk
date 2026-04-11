import io
from datetime import datetime, timezone

from geoalchemy2.elements import WKTElement
from openpyxl import load_workbook

from app.models.observation import Observation, ObservationStatus
from app.models.species import Species, SpeciesCategory, SpeciesGroup
from app.models.user import User, UserRole


def _create_user(db, *, external_id: str, role: UserRole = UserRole.employee) -> User:
    user = User(
        external_id=external_id,
        display_name=external_id,
        email=f"{external_id}@nlmk.com",
        role=role,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def _create_species(db, *, suffix: str, group: SpeciesGroup) -> Species:
    species = Species(
        name_ru=f"Species {suffix}",
        name_latin=f"Species {suffix} latin",
        group=group,
        category=SpeciesCategory.typical,
    )
    db.add(species)
    db.commit()
    db.refresh(species)
    return species


def _create_observation(
    db,
    *,
    author_id: int,
    species_id: int | None,
    group: str,
    status: ObservationStatus,
    is_incident: bool = False,
    comment: str = "",
) -> Observation:
    obs = Observation(
        author_id=author_id,
        species_id=species_id,
        group=group,
        observed_at=datetime(2026, 4, 11, 9, 15, tzinfo=timezone.utc),
        location_point=WKTElement("POINT(39.60 52.59)", srid=4326),
        status=status,
        comment=comment,
        is_incident=is_incident,
        safety_checked=True,
    )
    db.add(obs)
    db.commit()
    db.refresh(obs)
    return obs


def _open_export_workbook(content: bytes):
    return load_workbook(io.BytesIO(content))


def test_export_observations_requires_ecologist_or_admin(client, employee_token):
    response = client.get(
        "/api/export/observations",
        headers={"Authorization": f"Bearer {employee_token}"},
    )
    assert response.status_code == 403


def test_export_observations_applies_filters_and_formats_xlsx(client, db, ecologist_token):
    author = _create_user(db, external_id="export-author-001")
    bird_species = _create_species(db, suffix="Bird", group=SpeciesGroup.birds)
    plant_species = _create_species(db, suffix="Plant", group=SpeciesGroup.plants)

    included = _create_observation(
        db,
        author_id=author.id,
        species_id=bird_species.id,
        group=SpeciesGroup.birds.value,
        status=ObservationStatus.confirmed,
        is_incident=True,
        comment="included row",
    )
    _create_observation(
        db,
        author_id=author.id,
        species_id=plant_species.id,
        group=SpeciesGroup.plants.value,
        status=ObservationStatus.confirmed,
        is_incident=False,
        comment="filtered by group",
    )
    _create_observation(
        db,
        author_id=author.id,
        species_id=bird_species.id,
        group=SpeciesGroup.birds.value,
        status=ObservationStatus.on_review,
        is_incident=False,
        comment="filtered by status",
    )

    response = client.get(
        "/api/export/observations?group=birds&status=confirmed",
        headers={"Authorization": f"Bearer {ecologist_token}"},
    )
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=\"observations_" in response.headers["content-disposition"]
    assert response.headers["content-disposition"].endswith(".xlsx\"")

    workbook = _open_export_workbook(response.content)
    sheet = workbook["Наблюдения"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "ID",
        "Группа",
        "Дата",
        "Статус",
        "Вид",
        "Комментарий",
        "Широта",
        "Долгота",
        "Инцидент",
        "Создано",
    )
    assert len(rows) == 2

    exported = rows[1]
    assert exported[0] == included.id
    assert exported[1] == SpeciesGroup.birds.value
    assert exported[3] == ObservationStatus.confirmed.value
    assert exported[4] == f"{bird_species.name_ru} ({bird_species.name_latin})"
    assert exported[5] == "included row"
    assert exported[8] == "Да"


def test_export_observations_handles_missing_species_reference(client, db, ecologist_token):
    author = _create_user(db, external_id="export-author-002")
    observation = _create_observation(
        db,
        author_id=author.id,
        species_id=None,
        group=SpeciesGroup.mammals.value,
        status=ObservationStatus.confirmed,
        comment="no species linked",
    )

    response = client.get(
        "/api/export/observations?status=confirmed",
        headers={"Authorization": f"Bearer {ecologist_token}"},
    )
    assert response.status_code == 200

    workbook = _open_export_workbook(response.content)
    rows = list(workbook["Наблюдения"].iter_rows(values_only=True))
    exported = rows[1]
    assert exported[0] == observation.id
    assert exported[4] in ("", None)
    assert exported[5] == "no species linked"
